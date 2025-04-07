import pandas as pd
from openpyxl.styles import Alignment, numbers
def guardar_csv_datos(nombre_archivo, matriz):
    matriz.to_csv(nombre_archivo, index=False)
def guardar_excel_datos(nombre_archivo, matriz):
    with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
        matriz.to_excel(writer, index=False, sheet_name="Datos Formateados")
        workbook = writer.book
        sheet = writer.sheets["Datos Formateados"]
        # 🔹 Aplicar formato numérico con símbolo $ y alineación a la derecha
        formato_moneda = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Formato de moneda en Excel
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):  # Solo aplicar formato a valores numéricos
                    cell.number_format = formato_moneda  # Formato moneda con $
                    cell.alignment = Alignment(horizontal="right")  # Alinear a la derecha
def guardar_data_frame_a_csv_datos(nombre_archivo, registros_database):
    cabecera = ["ID"
            ,   "ID_PROYECTO_CONTRATO_CONVENIO"
            ,   "ESTADO"
            ,   "ENTIDAD_CONTRATANTE"
            ,   "VALOR_APORTE_DE_LA_ENTIDAD"
            ,   "PORCENTAJE_APORTE_DE_LA_ENTIDAD"
            ,   "VALOR_APORTE_FONPACIFICO"
            ,   "PORCENTAJE_APORTE_FONPACIFICO"
            ,   "ENTIDAD_DEL_DERIVADO"
            ,   "CODIGO_FACTURACION"
            ,   "FECHA_FACTURACION"
            ,   "VALOR_FACTURACION"
            ,   "VALOR_EN_SALDO"
            ,   "VALOR_DE_ABONOS"
            ,   "CODIGO_RECAUDO"
            ,   "FECHA_DE_RECAUDACION"
            ,   "VALOR_CONSIGNADO"
            ,   "VALOR_DEDUCCIONES"
            ,   "PORC_5_SOBRE_TOTAL_REGALIAS"
            ,   "PORC_5_SOBRE_VALOR_CONTRATO_SIN_APORTE"
            ,   "PORC_5_DEL_58_INTERADMINISTRATIVO"
            ,   "VALOR_APORTADO_CONTRATO_POR_FONPACI"
            ,   "POR_42_DEDUCIBLE_DE_INTERADMINSTRATIVOS"]
    df = pd.DataFrame(registros_database, columns=cabecera)
    df.to_csv(nombre_archivo, index=False)
def mostrar_datos_database(registros_database):
    for datos in registros_database:
        print(f"Datos DataBase Total ------> : {datos}")
def imprimir_datos_dataframe(df_combinado_facturacion_recaudos_matriz):
    for registro in df_combinado_facturacion_recaudos_matriz.to_dict(orient='records'):
        print(registro["ID PROYECTO / ID CONTRATO / ID CONVENIO"])
        print(registro)